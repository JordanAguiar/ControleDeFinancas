from openpyxl import load_workbook
import openpyxl
from pathlib import Path

ARQUIVO = "financeiro.xlsx"

COLUNAS_GASTOS  = ["Data", "Descrição", "Categoria", "Valor (R$)"]
COLUNAS_DIVIDAS = ["Credor", "Descrição", "Valor Total (R$)", "Valor Pago (R$)", "Vencimento", "Status"]
COLUNAS_RECEITAS = ["Data", "Descrição", "Categoria", "Valor (R$)"]


def inicializar_arquivo():
    if Path(ARQUIVO).exists():
        return
    wb = openpyxl.Workbook()

    ws_gastos = wb.active
    ws_gastos.title = "Gastos"
    ws_gastos.append(COLUNAS_GASTOS)

    wb.create_sheet("Dividas").append(COLUNAS_DIVIDAS)
    wb.create_sheet("Receitas").append(COLUNAS_RECEITAS)
    wb.create_sheet("Meta").append(["Meta Mensal (R$)", "Data Definida"])

    wb.save(ARQUIVO)


def abrir():
    return load_workbook(ARQUIVO)