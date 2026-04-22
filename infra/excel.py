from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
from core import (Gasto, Receita, Divida, Investimento,
                  GastoFixo, Parcela)

ARQUIVO = "financeiro.xlsx"

ABAS = {
    "Gastos":        ["Data", "Descrição", "Categoria", "Valor (R$)"],
    "Receitas":      ["Data", "Descrição", "Categoria", "Valor (R$)"],
    "Dividas":       ["Credor", "Descrição", "Valor Total (R$)",
                      "Valor Pago (R$)", "Vencimento", "Status"],
    "Investimentos": ["Data", "Descrição", "Categoria",
                      "Valor (R$)", "Instituição"],
    "Fixos":         ["Descrição", "Categoria", "Valor (R$)",
                      "Dia Vencimento", "Ativo"],
    "Parcelas":      ["Descrição", "Categoria", "Valor Total (R$)",
                      "Nº Parcelas", "Parcelas Pagas",
                      "Valor Parcela (R$)", "Data Início", "Ativo"],
    "Meta":          ["Meta Mensal (R$)", "Data Definida"],
}


def inicializar():
    """Cria o arquivo Excel com todas as abas se não existir."""
    if Path(ARQUIVO).exists():
        return
    wb = openpyxl.Workbook()
    primeira = True
    for nome, colunas in ABAS.items():
        if primeira:
            ws = wb.active
            ws.title = nome
            primeira = False
        else:
            ws = wb.create_sheet(nome)
        ws.append(colunas)
    wb.save(ARQUIVO)


def _abrir():
    return load_workbook(ARQUIVO)


def _garantir_aba(wb, nome):
    """Cria a aba se não existir."""
    if nome not in wb.sheetnames:
        ws = wb.create_sheet(nome)
        ws.append(ABAS[nome])
    return wb[nome]


# ── Gastos ───────────────────────────────────────────────────

def salvar_gasto(gasto: Gasto):
    wb = _abrir()
    _garantir_aba(wb, "Gastos").append(
        [gasto.data, gasto.descricao, gasto.categoria, gasto.valor])
    wb.save(ARQUIVO)


def buscar_gastos() -> list[Gasto]:
    wb = _abrir()
    ws = _garantir_aba(wb, "Gastos")
    return [
        Gasto(data=l[0], descricao=l[1], categoria=l[2], valor=l[3])
        for l in ws.iter_rows(min_row=2, values_only=True) if l[0]
    ]


def remover_gasto(indice: int):
    wb = _abrir()
    wb["Gastos"].delete_rows(indice + 2)
    wb.save(ARQUIVO)


# ── Receitas ─────────────────────────────────────────────────

def salvar_receita(receita: Receita):
    wb = _abrir()
    _garantir_aba(wb, "Receitas").append(
        [receita.data, receita.descricao, receita.categoria, receita.valor])
    wb.save(ARQUIVO)


def buscar_receitas() -> list[Receita]:
    wb = _abrir()
    ws = _garantir_aba(wb, "Receitas")
    return [
        Receita(data=l[0], descricao=l[1], categoria=l[2], valor=l[3])
        for l in ws.iter_rows(min_row=2, values_only=True) if l[0]
    ]


def remover_receita(indice: int):
    wb = _abrir()
    wb["Receitas"].delete_rows(indice + 2)
    wb.save(ARQUIVO)


# ── Dívidas ──────────────────────────────────────────────────

def salvar_divida(divida: Divida):
    wb = _abrir()
    _garantir_aba(wb, "Dividas").append([
        divida.credor, divida.descricao, divida.valor_total,
        divida.valor_pago, divida.vencimento, divida.status
    ])
    wb.save(ARQUIVO)


def buscar_dividas() -> list[Divida]:
    wb = _abrir()
    ws = _garantir_aba(wb, "Dividas")
    return [
        Divida(credor=l[0], descricao=l[1], valor_total=l[2],
               valor_pago=l[3], vencimento=l[4], status=l[5])
        for l in ws.iter_rows(min_row=2, values_only=True) if l[0]
    ]


def remover_divida(indice: int):
    wb = _abrir()
    wb["Dividas"].delete_rows(indice + 2)
    wb.save(ARQUIVO)


# ── Investimentos ────────────────────────────────────────────

def salvar_investimento(inv: Investimento):
    wb = _abrir()
    _garantir_aba(wb, "Investimentos").append(
        [inv.data, inv.descricao, inv.categoria,
         inv.valor, inv.instituicao])
    wb.save(ARQUIVO)


def buscar_investimentos() -> list[Investimento]:
    wb = _abrir()
    ws = _garantir_aba(wb, "Investimentos")
    return [
        Investimento(data=l[0], descricao=l[1], categoria=l[2],
                     valor=l[3], instituicao=l[4])
        for l in ws.iter_rows(min_row=2, values_only=True) if l[0]
    ]


def remover_investimento(indice: int):
    wb = _abrir()
    wb["Investimentos"].delete_rows(indice + 2)
    wb.save(ARQUIVO)


# ── Gastos Fixos ─────────────────────────────────────────────

def salvar_fixo(fixo: GastoFixo):
    wb = _abrir()
    _garantir_aba(wb, "Fixos").append([
        fixo.descricao, fixo.categoria,
        fixo.valor, fixo.dia_vencimento, fixo.ativo
    ])
    wb.save(ARQUIVO)


def buscar_fixos() -> list[GastoFixo]:
    wb = _abrir()
    ws = _garantir_aba(wb, "Fixos")
    return [
        GastoFixo(descricao=l[0], categoria=l[1], valor=l[2],
                  dia_vencimento=l[3], ativo=l[4])
        for l in ws.iter_rows(min_row=2, values_only=True) if l[0]
    ]


def remover_fixo(indice: int):
    wb = _abrir()
    wb["Fixos"].delete_rows(indice + 2)
    wb.save(ARQUIVO)


# ── Parcelas ─────────────────────────────────────────────────

def salvar_parcela(parcela: Parcela):
    wb = _abrir()
    _garantir_aba(wb, "Parcelas").append([
        parcela.descricao, parcela.categoria, parcela.valor_total,
        parcela.num_parcelas, parcela.parcelas_pagas,
        parcela.valor_parcela, parcela.data_inicio, parcela.ativo
    ])
    wb.save(ARQUIVO)


def buscar_parcelas() -> list[Parcela]:
    wb = _abrir()
    ws = _garantir_aba(wb, "Parcelas")
    return [
        Parcela(descricao=l[0], categoria=l[1], valor_total=l[2],
                num_parcelas=l[3], parcelas_pagas=l[4],
                valor_parcela=l[5], data_inicio=l[6], ativo=l[7])
        for l in ws.iter_rows(min_row=2, values_only=True) if l[0]
    ]


def atualizar_parcelas_pagas(indice: int, pagas: int):
    wb = _abrir()
    wb["Parcelas"].cell(row=indice + 2, column=5, value=pagas)
    wb.save(ARQUIVO)


def remover_parcela(indice: int):
    wb = _abrir()
    wb["Parcelas"].delete_rows(indice + 2)
    wb.save(ARQUIVO)


# ── Meta ─────────────────────────────────────────────────────

def salvar_meta(valor: float):
    wb = _abrir()
    ws = _garantir_aba(wb, "Meta")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None
    ws.cell(row=2, column=1, value=valor)
    ws.cell(row=2, column=2, value=datetime.now().strftime("%d/%m/%Y"))
    wb.save(ARQUIVO)


def buscar_meta() -> float:
    wb = _abrir()
    if "Meta" not in wb.sheetnames:
        return 0.0
    valor = wb["Meta"].cell(row=2, column=1).value
    return float(valor) if valor else 0.0